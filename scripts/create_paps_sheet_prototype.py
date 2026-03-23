from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT_PATH = Path("/Volumes/DATA/Dev/Codex/paps-tracker/output/spreadsheet/paps-google-sheet-prototype.xlsx")


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="EAF4FF")
STATIC_FILL = PatternFill("solid", fgColor="EFEFEF")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN = Side(style="thin", color="D9D9D9")


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=THIN)


def style_subheader(cell):
    cell.fill = SUBHEADER_FILL
    cell.font = BOLD_FONT
    cell.alignment = Alignment(vertical="center")


def autofit(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def create_settings_sheet(wb):
    ws = wb.active
    ws.title = "설정"

    ws["A1"] = "항목"
    ws["B1"] = "값"
    ws["C1"] = "설명"
    style_header(ws)

    rows = [
        ("학교명", "도촌초등학교", "교사가 관리 페이지에서 설정"),
        ("학년도", "2026", "모든 탭에서 학년도 컬럼과 함께 사용"),
        ("담당교사 이메일", "teacher@example.com", "구글 로그인 계정"),
        ("기본 세션 유형", "연습 기록", "세션 생성 시 바꿀 수 있음"),
        ("입력 화면 유형", "1반형 / 2반 분할형", "관리 페이지에서 선택"),
        ("2반 분할 규칙", "같은 종목만 동시 기록", "사용자 승인 반영"),
        ("학생 조회 정책", "제출 직후에만 자기 기록 확인", "공용 기기 보호 정책"),
        ("시트 템플릿 버전", "v0.1-prototype", "프로토타입 예시"),
        ("기록 보관 정책", "최소 해당 학년도 보관", "이전 학년도는 조회용 유지 또는 별도 백업"),
        ("템플릿 안내 링크", "https://docs.google.com/spreadsheets/", "실제 구현 시 복사 링크 버튼으로 연결"),
    ]

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws[f"A{row_idx}"].fill = STATIC_FILL
        ws[f"A{row_idx}"].font = BOLD_FONT
        ws[f"B{row_idx}"].fill = INPUT_FILL

    ws["E1"] = "사용 탭"
    ws["F1"] = "역할"
    for cell in ws["E1:F1"][0]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")

    tab_rows = [
        ("학생명단", "학생 기본 정보와 활성 여부"),
        ("세션기록", "학생 입력 원본을 시도별로 모두 누적 저장"),
        ("학생요약", "대표값 기준 최신값, 최고값, 변화량"),
        ("공식평가요약", "공식 평가 세션 대표값과 등급"),
        ("오류로그", "시트 동기화 실패와 주소 오류"),
        ("수정로그", "교사 대표값 선택 및 수정 이력"),
    ]
    for row_idx, row in enumerate(tab_rows, start=2):
        for col_idx, value in enumerate(row, start=5):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    autofit(ws, [18, 30, 44, 4, 20, 36])


def create_students_sheet(wb):
    ws = wb.create_sheet("학생명단")
    headers = ["학생ID", "학년도", "학년", "반", "번호", "이름", "성별", "활성", "비고"]
    ws.append(headers)
    style_header(ws)

    rows = [
        ["S-2026-5-2-01", 2026, 5, 2, 1, "김민준", "남", "Y", ""],
        ["S-2026-5-2-02", 2026, 5, 2, 2, "박서연", "여", "Y", ""],
        ["S-2026-5-2-03", 2026, 5, 2, 3, "이도윤", "남", "Y", ""],
        ["S-2026-5-3-01", 2026, 5, 3, 1, "최하은", "여", "Y", ""],
        ["S-2026-5-3-02", 2026, 5, 3, 2, "정예준", "남", "Y", ""],
        ["S-2026-5-3-03", 2026, 5, 3, 3, "윤지아", "여", "Y", ""],
    ]
    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=THIN)
        row[0].fill = GOOD_FILL
        row[7].fill = INPUT_FILL

    freeze_and_filter(ws)
    autofit(ws, [18, 10, 8, 8, 8, 12, 10, 10, 24])


def create_records_sheet(wb):
    ws = wb.create_sheet("세션기록")
    headers = [
        "기록ID",
        "세션ID",
        "세션명",
        "학년도",
        "측정일",
        "세션유형",
        "입력화면유형",
        "대상반표시",
        "실제반",
        "종목",
        "단위",
        "학생ID",
        "학생이름",
        "시도순번",
        "원측정값",
        "대표값선택",
        "대표값선정교사",
        "공식등급",
        "제출시각",
        "동기화상태",
        "비고",
    ]
    ws.append(headers)
    style_header(ws)

    rows = [
        ["R-0001", "SES-2026-05-2-3-0601", "5학년 2반+3반 6월 연습 제자리멀리뛰기", 2026, "2026-06-01", "연습", "2반 분할형", "2반+3반", 2, "제자리멀리뛰기", "cm", "S-2026-5-2-01", "김민준", 1, 145, "N", "", "", "2026-06-01 09:03:11", "완료", ""],
        ["R-0002", "SES-2026-05-2-3-0601", "5학년 2반+3반 6월 연습 제자리멀리뛰기", 2026, "2026-06-01", "연습", "2반 분할형", "2반+3반", 2, "제자리멀리뛰기", "cm", "S-2026-5-2-01", "김민준", 2, 152, "Y", "teacher@example.com", "", "2026-06-01 09:04:33", "완료", "교사가 대표값 선택"],
        ["R-0003", "SES-2026-05-2-3-0601", "5학년 2반+3반 6월 연습 제자리멀리뛰기", 2026, "2026-06-01", "연습", "2반 분할형", "2반+3반", 3, "제자리멀리뛰기", "cm", "S-2026-5-3-01", "최하은", 1, 138, "Y", "teacher@example.com", "", "2026-06-01 09:05:09", "완료", ""],
        ["R-0004", "SES-2026-05-2-0615", "5학년 2반 6월 공식 제자리멀리뛰기", 2026, "2026-06-15", "공식", "1반형", "2반", 2, "제자리멀리뛰기", "cm", "S-2026-5-2-02", "박서연", 1, 149, "Y", "teacher@example.com", 3, "2026-06-15 10:11:52", "완료", ""],
        ["R-0005", "SES-2026-05-2-0615", "5학년 2반 6월 공식 제자리멀리뛰기", 2026, "2026-06-15", "공식", "1반형", "2반", 2, "제자리멀리뛰기", "cm", "S-2026-5-2-03", "이도윤", 1, 161, "N", "", 2, "2026-06-15 10:13:04", "완료", ""],
        ["R-0006", "SES-2026-05-2-0615", "5학년 2반 6월 공식 제자리멀리뛰기", 2026, "2026-06-15", "공식", "1반형", "2반", 2, "제자리멀리뛰기", "cm", "S-2026-5-2-03", "이도윤", 2, 166, "Y", "teacher@example.com", 2, "2026-06-15 10:14:28", "완료", "2차 시도 대표값"],
    ]
    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[14].fill = INPUT_FILL
        row[15].fill = WARN_FILL if row[15].value == "N" else GOOD_FILL
        row[19].fill = GOOD_FILL if row[19].value == "완료" else ERROR_FILL

    freeze_and_filter(ws)
    autofit(ws, [12, 18, 28, 10, 12, 12, 14, 12, 8, 16, 8, 18, 12, 10, 12, 12, 18, 10, 18, 10, 24])


def create_summary_sheet(wb):
    ws = wb.create_sheet("학생요약")
    headers = [
        "학생ID",
        "이름",
        "학년",
        "반",
        "종목",
        "최신대표값",
        "단위",
        "직전대표값",
        "변화량",
        "최고대표값",
        "최근측정일",
        "학생표시문구",
    ]
    ws.append(headers)
    style_header(ws)

    sample_rows = [
        ["S-2026-5-2-01", "김민준", 5, 2, "제자리멀리뛰기", 152, "cm", 145, "=F2-H2", 152, "2026-06-01", "지난 기록 대비 "],
        ["S-2026-5-3-01", "최하은", 5, 3, "제자리멀리뛰기", 138, "cm", "", '=IF(H3="","",F3-H3)', 138, "2026-06-01", "첫 기록"],
        ["S-2026-5-2-02", "박서연", 5, 2, "제자리멀리뛰기", 149, "cm", "", '=IF(H4="","",F4-H4)', 149, "2026-06-15", "공식 기록 완료"],
    ]
    for row in sample_rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=THIN)
        row[5].fill = INPUT_FILL
        row[8].fill = SUBHEADER_FILL

    freeze_and_filter(ws)
    autofit(ws, [18, 12, 8, 8, 16, 12, 8, 12, 10, 12, 12, 22])


def create_official_sheet(wb):
    ws = wb.create_sheet("공식평가요약")
    headers = [
        "학생ID",
        "이름",
        "학년",
        "반",
        "종목",
        "대표값",
        "단위",
        "공식등급",
        "측정일",
        "세션명",
        "비고",
    ]
    ws.append(headers)
    style_header(ws)

    rows = [
        ["S-2026-5-2-02", "박서연", 5, 2, "제자리멀리뛰기", 149, "cm", 3, "2026-06-15", "5학년 2반 6월 공식 제자리멀리뛰기", ""],
        ["S-2026-5-2-03", "이도윤", 5, 2, "제자리멀리뛰기", 166, "cm", 2, "2026-06-15", "5학년 2반 6월 공식 제자리멀리뛰기", "2차 시도 대표값"],
    ]
    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=THIN)
        row[7].fill = WARN_FILL

    freeze_and_filter(ws)
    autofit(ws, [18, 12, 8, 8, 16, 10, 8, 10, 12, 28, 20])


def create_error_sheet(wb):
    ws = wb.create_sheet("오류로그")
    headers = ["시간", "수준", "구분", "메시지", "관련ID", "재시도상태", "해결시각"]
    ws.append(headers)
    style_header(ws)

    rows = [
        ["2026-06-15 10:15:01", "WARN", "시트동기화", "Google Sheets 쓰기 지연으로 재시도 대기", "R-0006", "대기", ""],
        ["2026-06-15 10:16:44", "INFO", "시트동기화", "재동기화 성공", "R-0006", "완료", "2026-06-15 10:16:44"],
    ]
    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[1].fill = WARN_FILL if row[1].value == "WARN" else GOOD_FILL

    freeze_and_filter(ws)
    autofit(ws, [18, 10, 12, 38, 12, 12, 18])


def create_audit_sheet(wb):
    ws = wb.create_sheet("수정로그")
    headers = ["시간", "교사계정", "세션ID", "학생ID", "종목", "작업", "이전기록ID", "선택기록ID", "사유"]
    ws.append(headers)
    style_header(ws)

    rows = [
        ["2026-06-01 09:05:00", "teacher@example.com", "SES-2026-05-2-3-0601", "S-2026-5-2-01", "제자리멀리뛰기", "대표값선택", "R-0001", "R-0002", "2차 측정값 채택"],
        ["2026-06-15 10:14:40", "teacher@example.com", "SES-2026-05-2-0615", "S-2026-5-2-03", "제자리멀리뛰기", "대표값선택", "R-0005", "R-0006", "2차 시도 기록 채택"],
    ]
    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    freeze_and_filter(ws)
    autofit(ws, [18, 24, 18, 18, 16, 14, 12, 12, 22])


def add_notes(ws):
    ws["A10"] = "입력 규칙 메모"
    style_subheader(ws["A10"])
    ws["A11"] = "파란색 셀은 교사/앱 입력값, 회색은 정적 설명, 노란색은 검토가 필요한 값입니다."
    ws["A12"] = "실제 앱은 학생 입력을 세션기록 탭에 누적하고, 학생요약/공식평가요약 탭을 다시 계산해 씁니다."
    ws["A13"] = "구글 시트 프로토타입이므로 수식 결과는 Google Sheets에서 열면 자동 계산됩니다."
    ws.merge_cells("A11:C11")
    ws.merge_cells("A12:C12")
    ws.merge_cells("A13:C13")
    for row in (11, 12, 13):
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)


def main():
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    create_settings_sheet(wb)
    create_students_sheet(wb)
    create_records_sheet(wb)
    create_summary_sheet(wb)
    create_official_sheet(wb)
    create_error_sheet(wb)
    create_audit_sheet(wb)
    add_notes(wb["설정"])

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
